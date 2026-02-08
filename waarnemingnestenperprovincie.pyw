import tkinter as tk
from tkinter import ttk
from calendar import monthrange
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
import re
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import requests
from datetime import datetime, timedelta

# 📅 GUI voor maand-, jaar- en provincie-selectie
def start_scraping():
    global eerste_dag, laatste_dag, provincie_code, provincie_naam

    geselecteerde_index = maand_dropdown.current()
    geselecteerd_jaar = int(jaar_entry.get())

    if geselecteerde_index == 0:
        # Alle maanden
        eerste_dag = f"{geselecteerd_jaar}-01-01"
        laatste_dag = f"{geselecteerd_jaar}-12-31"
    else:
        geselecteerde_maand = geselecteerde_index  # 1 = januari
        eerste_dag = f"{geselecteerd_jaar}-{geselecteerde_maand:02d}-01"
        laatste_dag = f"{geselecteerd_jaar}-{geselecteerde_maand:02d}-{monthrange(geselecteerd_jaar, geselecteerde_maand)[1]}"

    provincie_naam = provincie_dropdown.get()
    provincie_code = provincie_dict[provincie_naam]
    root.destroy()

root = tk.Tk()
root.title("Waarnemingen ophalen")

# Popup centreren
window_width = 300
window_height = 200
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
x_cordinate = int((screen_width/2) - (window_width/2))
y_cordinate = int((screen_height/2) - (window_height/2))
root.geometry(f"{window_width}x{window_height}+{x_cordinate}+{y_cordinate}")

# Provincie dropdown
tk.Label(root, text="Kies een provincie:").grid(row=0, column=0, padx=10, pady=10)
provincies = [
    "Drenthe", "Flevoland", "Friesland", "Gelderland", "Groningen",
    "Limburg", "Noord-Brabant", "Noord-Holland", "Overijssel",
    "Utrecht", "Zeeland", "Zuid-Holland"
]
provincie_dict = {
    "Drenthe": 5, "Flevoland": 8, "Friesland": 3, "Gelderland": 7, "Groningen": 4,
    "Limburg": 11, "Noord-Brabant": 10, "Noord-Holland": 2, "Overijssel": 6,
    "Utrecht": 1, "Zeeland": 12, "Zuid-Holland": 9
}
provincie_dropdown = ttk.Combobox(root, values=provincies, state="readonly")
provincie_dropdown.grid(row=0, column=1, padx=10, pady=10)
provincie_dropdown.set("Utrecht")

# Maand en jaar
tk.Label(root, text="Kies een maand:").grid(row=1, column=0, padx=10, pady=10)
maanden = [
    "Alle maanden",
    "Januari", "Februari", "Maart", "April", "Mei", "Juni",
    "Juli", "Augustus", "September", "Oktober", "November", "December"
]

maand_dropdown = ttk.Combobox(root, values=maanden, state="readonly")
maand_dropdown.grid(row=1, column=1, padx=10, pady=10)

# Fix voor Tkinter-bug: eerst naar 0, dan naar huidige maand
maand_dropdown.current(0)
huidige_maand = datetime.now().month
maand_dropdown.current(huidige_maand)


tk.Label(root, text="Voer het jaar in:").grid(row=2, column=0, padx=10, pady=10)
jaar_entry = tk.Entry(root)
jaar_entry.grid(row=2, column=1, padx=10, pady=10)
jaar_entry.insert(0, str(datetime.now().year))

start_button = tk.Button(root, text="Start", command=start_scraping)
start_button.grid(row=3, column=0, columnspan=2, pady=20)

root.mainloop()

# Chrome-opties
options = Options()
options.add_argument("--start-maximized")
options.add_argument("--disable-notifications")
options.add_argument("--disable-background-networking")
options.add_argument("--disable-default-apps")
options.add_argument("--disable-popup-blocking")
options.add_argument("--remote-allow-origins=*")

def start_driver():
    return webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

driver_main = start_driver()
driver_detail = start_driver()
print("Browsers gestart")

# Reverse geocoding
geocode_cache = {}

def reverse_geocode(gps_str, delay=1.0):
    match = re.search(r'GPS\s*([\d.]+),\s*([\d.]+)', gps_str)
    if not match:
        return "Onbekend adres", "Onbekende gemeente"

    lat, lon = float(match.group(1)), float(match.group(2))

    # Cache lookup
    if (lat, lon) in geocode_cache:
        return geocode_cache[(lat, lon)]

    url = "https://nominatim.openstreetmap.org/reverse"
    params = {"lat": lat, "lon": lon, "format": "jsonv2", "addressdetails": 1}
    headers = {"User-Agent": "NestMonitorScript/1.0"}

    try:
        response = requests.get(url, params=params, headers=headers)
        response.raise_for_status()
        data = response.json()

        adres = data.get("display_name", "Onbekend adres")

        # Gemeente ophalen (Nominatim gebruikt soms municipality, soms city)
        gemeente = (
            data.get("address", {}).get("municipality")
            or data.get("address", {}).get("city")
            or data.get("address", {}).get("town")
            or data.get("address", {}).get("village")
            or "Onbekende gemeente"
        )

        geocode_cache[(lat, lon)] = (adres, gemeente)
        return adres, gemeente

    except:
        return "Geocode mislukt", "Onbekende gemeente"

    finally:
        time.sleep(delay)

# Eerder opgeslagen data
excel_path = f"aziatische_hoornaar_nesten_{provincie_naam.lower().replace('-', '_')}.xlsx"

# Bestaande data inlezen met behoud van URL uit hyperlinkformule
waarnemingen = []
bekende_coords = set()
bestaande_ids = set()

try:
    wb_existing = load_workbook(excel_path, data_only=False)
    ws_existing = wb_existing.active

    rows = list(ws_existing.iter_rows(values_only=False))
    headers = [cell.value for cell in rows[0]]

    link_col = headers.index("Link") if "Link" in headers else None
    gps_col = headers.index("GPS") if "GPS" in headers else None
    id_col = headers.index("Waarneming ID") if "Waarneming ID" in headers else None

    data = []
    for row in rows[1:]:
        values = []
        for i, cell in enumerate(row):
            val = cell.value
            if link_col is not None and i == link_col and isinstance(val, str) and val.startswith("=HYPERLINK("):
                m = re.search(r'HYPERLINK\("([^"]+)"', val)
                url = m.group(1) if m else ""
                values.append(url)
            else:
                values.append(val)
        data.append(values)

        if gps_col is not None:
            gps_val = row[gps_col].value
            if isinstance(gps_val, str):
                match = re.search(r'GPS\s*([\d.]+),\s*([\d.]+)', gps_val)
                if match:
                    lat = float(match.group(1))
                    lon = float(match.group(2))
                    bekende_coords.add((lat, lon))

    bestaand_df = pd.DataFrame(data, columns=headers)
    if "Waarneming ID" in bestaand_df.columns:
        bestaande_ids = set(bestaand_df["Waarneming ID"].astype(str))
    else:
        bestaande_ids = set()

except FileNotFoundError:
    bestaand_df = pd.DataFrame()
    bestaande_ids = set()

# verwerkte_ids voor deze run (voorkomt dubbele verwerking binnen blokken)
verwerkte_ids = set()

def is_doublure(gps_str, bekende_coords, tolerance=0.0005):
    match = re.search(r'GPS\s*([\d.]+),\s*([\d.]+)', gps_str)
    if not match:
        return False
    lat = float(match.group(1))
    lon = float(match.group(2))
    for known_lat, known_lon in bekende_coords:
        if abs(lat - known_lat) <= tolerance and abs(lon - known_lon) <= tolerance:
            return True
    bekende_coords.add((lat, lon))
    return False

sessie_teller = 0
max_per_sessie = 400

eerste_dag_dt = datetime.strptime(eerste_dag, "%Y-%m-%d")
laatste_dag_dt = datetime.strptime(laatste_dag, "%Y-%m-%d")

blok_start = eerste_dag_dt

while blok_start <= laatste_dag_dt:

    blok_einde = min(blok_start + timedelta(days=10), laatste_dag_dt)
    print(f"\n=== Blok {blok_start.date()} t/m {blok_einde.date()} ===")

    page = 1

    while True:

        if sessie_teller >= max_per_sessie:
            print(f"{sessie_teller} waarnemingen verwerkt — herstart browsers...")

            driver_main.quit()
            driver_detail.quit()
            time.sleep(2)

            driver_main = start_driver()
            driver_detail = start_driver()

            print("Nieuwe browsers gestart")
            sessie_teller = 0
            continue

        target_url = (
            "https://waarneming.nl/species/8807/observations/"
            f"?date_after={blok_start.date()}"
            f"&date_before={blok_einde.date()}"
            f"&country_division={provincie_code}"
            "&search=&advanced=on"
            "&activity=NEST"
            f"&page={page}"
        )

        print(f"\nPagina {page}: {target_url}")
        driver_main.get(target_url)

        try:
            WebDriverWait(driver_main, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "table.table tbody tr"))
            )
        except:
            print("Geen tabel gevonden — einde van dit blok.")
            break

        rows = driver_main.find_elements(By.CSS_SELECTOR, "table.table tbody tr")
        if not rows:
            print("Geen resultaten meer in dit blok.")
            break

        herhaling_detected = True

        for row in rows:

            sessie_teller += 1

            cols = row.find_elements(By.TAG_NAME, "td")
            if len(cols) >= 5:
                try:
                    link_element = cols[0].find_element(By.TAG_NAME, "a")
                    waarneming_url = link_element.get_attribute("href")
                    match = re.search(r'/observation/(\d+)', waarneming_url)
                    waarneming_id = match.group(1) if match else "Onbekend"
                except:
                    waarneming_url = ""
                    waarneming_id = "Onbekend"

                # Skip als deze waarneming al in Excel staat
                if waarneming_id in bestaande_ids:
                    continue

                # Skip als deze waarneming in deze run al is gezien
                if waarneming_id in verwerkte_ids:
                    continue

                verwerkte_ids.add(waarneming_id)
                herhaling_detected = False

                aantal = cols[1].text.strip()

                gps_raw = ""
                adres = "Onbekend"

                if waarneming_url:
                    try:
                        driver_detail.get(waarneming_url)
                        WebDriverWait(driver_detail, 10).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, "span.teramap-coordinates:nth-child(1)"))
                        )
                        gps_element = driver_detail.find_element(By.CSS_SELECTOR, "span.teramap-coordinates:nth-child(1)")
                        gps_raw = gps_element.text.strip()
                        adres, gemeente = reverse_geocode(gps_raw)
                    except:
                        print(f"GPS niet gevonden voor {waarneming_id}")

                doublure_status = is_doublure(gps_raw, bekende_coords)

                waarneming = {
                    "Waarneming ID": waarneming_id,
                    "Datum": cols[0].text.strip(),
                    "Aantal": aantal,
                    "Locatie": cols[2].text.strip(),
                    "Waarnemer": cols[3].text.strip(),
                    "Link": waarneming_url or "",
                    "GPS": gps_raw,
                    "Doublure": doublure_status,
                    "Adres": adres,
                    "Gemeente": gemeente
                }

                waarnemingen.append(waarneming)
                time.sleep(1)

        if herhaling_detected:
            print("Herhaling — einde van dit blok.")
            break

        page += 1
        time.sleep(2)

    blok_start = blok_einde + timedelta(days=1)

# Sluit browsers
driver_main.quit()
driver_detail.quit()

# 🧹 Nieuwe waarnemingen verwerken
nieuwe_df = pd.DataFrame(waarnemingen)

# Als er geen nieuwe waarnemingen zijn → lege df laten
if nieuwe_df.empty:
    nieuwe_df = pd.DataFrame()
else:
    if bestaande_ids:
        nieuwe_df = nieuwe_df[~nieuwe_df["Waarneming ID"].astype(str).isin(bestaande_ids)]

# Samenvoegen met bestaande data
samengevoegd_df = pd.concat([bestaand_df, nieuwe_df], ignore_index=True)

# Sorteren op Waarneming ID
if "Waarneming ID" in samengevoegd_df.columns:
    samengevoegd_df["Waarneming ID"] = pd.to_numeric(
        samengevoegd_df["Waarneming ID"], errors="coerce"
    )
    samengevoegd_df = samengevoegd_df.sort_values(by="Waarneming ID")

# 🔗 Hyperlinkfunctie (met komma)
def maak_hyperlink(url):
    if pd.isna(url) or url == "":
        return ""
    return f'=HYPERLINK("{url}", "Bekijk")'

if "Link" in samengevoegd_df.columns:
    samengevoegd_df["Link"] = samengevoegd_df["Link"].apply(maak_hyperlink)

# Wegschrijven naar Excel
samengevoegd_df.to_excel(excel_path, index=False)

# 🎨 Visuele markering van doublures
wb = load_workbook(excel_path)
ws = wb.active
fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

doublure_col = None
for i, cell in enumerate(ws[1], start=1):
    if cell.value == "Doublure":
        doublure_col = i
        break

if doublure_col:
    for row in ws.iter_rows(min_row=2):
        cell = row[doublure_col - 1]
        if cell.value is True:
            for c in row:
                c.fill = fill

wb.save(excel_path)
print("📁 Bestand aangevuld en doublures gemarkeerd.")

